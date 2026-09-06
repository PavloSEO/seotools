"""Regression for #491: Technologies sheet formula injection.

A crawled site fully controls the detected-technology ``name`` (sourced from
its own X-Powered-By response header, see seohead/recon/tech.py). Writing it
unescaped into an XLSX cell lets the audited site plant a live Excel formula
that executes the moment the operator opens the report (CWE-1236), the same
class #153 already fixed for the Findings and Pages sheets.

Synthetic fixtures only; no network access, no real credentials.
"""

from __future__ import annotations

from openpyxl import load_workbook

from seohead.audit.site import SCHEMA
from seohead.reports import build_report

BASE_DOCUMENT = {
    "schema": SCHEMA,
    "domain": "example.com",
    "url": "https://example.com/",
    "generated_at": "2026-09-05T00:00:00Z",
    "findings": [],
    "pages": [],
    "summary": {"pages_checked": 0, "findings_total": 0, "findings_by_severity": {}},
}


def _build(tmp_path, technologies):
    doc = {
        **BASE_DOCUMENT,
        "site": {"site": {}, "tech_detect": {"technologies": technologies}},
    }
    out = tmp_path / "tech.xlsx"
    result = build_report(doc, "xlsx", str(out))
    assert result["ok"] is True
    wb = load_workbook(out)
    return wb["Technologies"]


def test_formula_leading_tech_name_is_neutralised(tmp_path):
    """A crawled site's X-Powered-By-derived name is written inert, not as a live formula."""
    ws = _build(
        tmp_path,
        [
            {
                "category": "runtime",
                "name": "=cmd|' /C calc'!A1",
                "evidence": "x-powered-by: =cmd|' /C calc'!A1",
            }
        ],
    )
    name_cell = ws.cell(row=2, column=2)
    evidence_cell = ws.cell(row=2, column=3)

    assert name_cell.data_type == "s"
    assert evidence_cell.data_type == "s"
    # Still legible: the leading apostrophe strips off, the payload text survives.
    assert name_cell.value.lstrip("'") == "=cmd|' /C calc'!A1"
    assert evidence_cell.value.lstrip("'") == "x-powered-by: =cmd|' /C calc'!A1"


def test_ordinary_tech_name_is_unchanged(tmp_path):
    """Negative control: a normal name renders byte for byte, no apostrophe added."""
    ws = _build(
        tmp_path,
        [{"category": "framework", "name": "Next.js", "evidence": "detected via build id"}],
    )
    name_cell = ws.cell(row=2, column=2)
    evidence_cell = ws.cell(row=2, column=3)

    assert name_cell.data_type == "s"
    assert name_cell.value == "Next.js"
    assert evidence_cell.value == "detected via build id"
