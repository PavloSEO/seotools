"""Bulk audit and report tests that validate orchestration without network access."""

from __future__ import annotations

import json

from seohead.audit.site import (
    PAGE_TOOLS,
    SCHEMA,
    SEVERITY_RULES,
    SITE_TOOLS,
    _first_h1,
    _page_row,
    _urls_from_sitemap,
    audit_site,
    classify,
)
from seohead.reports import FORMATS, build_report
from seohead.servers import handlers

# ── finding severity ─────────────────────────────────────────────────────────


def test_severity_rules_are_ordered_from_worst():
    """The first matching rule wins, so critical rules must precede warnings."""
    levels = [level for _, level in SEVERITY_RULES]
    assert levels == sorted(levels, key=lambda x: {"critical": 0, "warning": 1}.get(x, 2))


def test_page_that_robots_see_empty_is_critical():
    assert classify('raw HTML contains an empty container <div id="root">') == "critical"
    assert classify("the crawler receives an empty page") == "critical"


def test_self_removal_from_index_is_critical():
    assert classify("regional pages are canonicalized to another host") == "critical"
    assert classify("regional pages use noindex") == "critical"


def test_ordinary_observation_is_a_notice():
    assert classify("brotli is not enabled") == "notice"
    assert classify("") == "notice"


# ── page list ────────────────────────────────────────────────────────────────


def test_urls_are_read_from_the_sitemap_records():
    sitemap = {
        "count": 3,
        "urls": [
            {"loc": "https://example.com/a", "lastmod": "2026-01-01"},
            {"loc": "https://example.com/b"},
            {"loc": "https://example.com/a"},  # duplicate
        ],
    }
    assert _urls_from_sitemap(sitemap, 10) == [
        "https://example.com/a",
        "https://example.com/b",
    ]


def test_plain_string_urls_work_too():
    assert _urls_from_sitemap({"urls": ["https://example.com/x"]}, 5) == ["https://example.com/x"]


def test_limit_is_respected():
    sitemap = {"urls": [{"loc": f"https://example.com/{i}"} for i in range(100)]}
    assert len(_urls_from_sitemap(sitemap, 7)) == 7


def test_nested_shapes_are_still_walked():
    """A response-shape change must not silently reduce the audit to zero URLs."""
    weird = {"sitemaps": [{"entries": [{"loc": "https://example.com/deep"}]}]}
    assert _urls_from_sitemap(weird, 5) == ["https://example.com/deep"]


# ── multiple declared sitemaps (#200) ────────────────────────────────────────

FIRST_SITEMAP = "https://example.test/sitemap-pages.xml"
SECOND_SITEMAP = "https://example.test/sitemap-products.xml"


def test_every_declared_sitemap_is_sampled_not_just_the_first(monkeypatch):
    """robots.txt can declare more than one independent Sitemap: directive; a page
    only that second sitemap lists must still reach page selection, not be
    silently dropped because ``site.py`` only ever fetched the first."""
    calls: list[str] = []

    def fake_robots_check(url: str) -> dict:
        return {"ok": True, "sitemaps": [FIRST_SITEMAP, SECOND_SITEMAP]}

    def fake_sitemap_crawl(url: str | None = None, **_kw: object) -> dict:
        calls.append(url or "")
        loc = (
            "https://example.test/page-a"
            if url == FIRST_SITEMAP
            else "https://example.test/product-b"
        )
        return {"ok": True, "urls": [{"loc": loc}]}

    # audit_site is handed the tools it composes rather than importing the server
    # layer to fetch them (#221), so the stub goes in through that seam instead of
    # being patched onto handlers -- which is the seam existing for exactly this.
    tools = {
        **handlers.HANDLERS,
        "robots_check": fake_robots_check,
        "sitemap_crawl": fake_sitemap_crawl,
    }

    skip = [t for t in SITE_TOOLS if t not in {"robots_check", "sitemap_crawl"}] + list(PAGE_TOOLS)
    result = audit_site("https://example.test/", skip=skip, tools=tools)

    assert calls == [FIRST_SITEMAP, SECOND_SITEMAP]
    selected = {page["url"] for page in result["pages"]}
    assert "https://example.test/page-a" in selected
    assert "https://example.test/product-b" in selected


def _multi_root_audit(sitemap_crawl):
    """Run audit_site with two declared roots and a caller-supplied sitemap_crawl stub."""

    def fake_robots_check(url: str) -> dict:
        return {"ok": True, "sitemaps": [FIRST_SITEMAP, SECOND_SITEMAP]}

    tools = {
        **handlers.HANDLERS,
        "robots_check": fake_robots_check,
        "sitemap_crawl": sitemap_crawl,
    }
    skip = [t for t in SITE_TOOLS if t not in {"robots_check", "sitemap_crawl"}] + list(PAGE_TOOLS)
    return audit_site("https://example.test/", skip=skip, tools=tools)


def test_two_partial_roots_are_not_reported_clean():
    """Two roots that both fail must not collapse into ``{"ok": True, "urls": []}`` —
    that is the exact aggregate the projection loss (#310) produced."""

    def fake_sitemap_crawl(url: str | None = None, **_kw: object) -> dict:
        return {
            "ok": True,
            "urls": [],
            "errors": [{"url": url, "error": "synthetic child unavailable"}],
            "truncated": True,
        }

    result = _multi_root_audit(fake_sitemap_crawl)
    sitemap = result["site"]["sitemap_crawl"]

    assert sitemap["ok"] is False
    assert len(sitemap["errors"]) == 2
    assert sitemap["truncated"] is True
    assert result["summary"]["tools_failed"], "a totally failed multi-root sitemap must be reported"


def test_one_good_root_keeps_urls_and_names_the_failed_root():
    def fake_sitemap_crawl(url: str | None = None, **_kw: object) -> dict:
        if url == FIRST_SITEMAP:
            return {"ok": True, "urls": [{"loc": "https://example.test/page-a"}]}
        return {"ok": True, "urls": [], "errors": [{"url": url, "error": "404"}]}

    result = _multi_root_audit(fake_sitemap_crawl)
    sitemap = result["site"]["sitemap_crawl"]

    assert sitemap["ok"] is True
    assert {e["loc"] for e in sitemap["urls"]} == {"https://example.test/page-a"}
    assert sitemap["errors"] == [{"url": SECOND_SITEMAP, "error": "404"}]
    # A non-fatal partial result still has to surface in the document, not only in
    # the raw site payload -- findings feed the Markdown/DOCX/XLSX report writers.
    named = [f for f in result["findings"] if SECOND_SITEMAP in f["text"]]
    assert named
    # And it has to surface at the severity that keeps it visible. classify()
    # matches SEVERITY_RULES by plain substring, so a finding whose wording drifts
    # by one word silently becomes a notice and sorts to the bottom of every
    # report -- under exactly the partial evidence it was written to raise.
    assert [f["severity"] for f in named] == ["critical"]


def test_all_successful_multi_root_shape_is_unchanged():
    """The #200 all-successful aggregate keeps its original shape: no ``errors`` or
    ``truncated`` keys appear when nothing went wrong."""

    def fake_sitemap_crawl(url: str | None = None, **_kw: object) -> dict:
        loc = (
            "https://example.test/page-a"
            if url == FIRST_SITEMAP
            else "https://example.test/product-b"
        )
        return {"ok": True, "urls": [{"loc": loc}]}

    result = _multi_root_audit(fake_sitemap_crawl)
    sitemap = result["site"]["sitemap_crawl"]

    assert sitemap == {
        "ok": True,
        "urls": [
            {"loc": "https://example.test/page-a"},
            {"loc": "https://example.test/product-b"},
        ],
        "sources": [FIRST_SITEMAP, SECOND_SITEMAP],
    }


# ── page row ─────────────────────────────────────────────────────────────────

PARSE_RESULT = {
    "count": 1,
    "results": [
        {
            "url": "https://example.com/p",
            "status_code": 200,
            "title": "Page title",
            "meta_description": "Page description",
            "canonical": "https://example.com/p",
            "word_count": 640,
            "headings": {"h1": ["Primary H1"]},
        }
    ],
}


def test_page_row_reads_the_batch_shape_of_parse():
    """Parse returns a batch shape; reading its top level would empty report columns."""
    row = _page_row("https://example.com/p", {"parse": PARSE_RESULT})
    assert row["status"] == 200
    assert row["words"] == 640
    assert row["title_length"] == len("Page title")
    assert row["description_length"] == len("Page description")
    assert row["h1"] == "Primary H1"


def test_h1_is_found_in_any_of_the_known_shapes():
    assert _first_h1({"h1": "direct value"}) == "direct value"
    assert _first_h1({"headings": {"h1": ["from mapping"]}}) == "from mapping"
    assert _first_h1({"headings": [{"level": "h1", "text": "from list"}]}) == "from list"
    assert _first_h1({}) == ""


def test_failed_tool_becomes_a_page_issue_not_silence():
    row = _page_row(
        "https://example.com/p",
        {"schema_check": {"ok": False, "error": "request timed out"}},
    )
    assert any("request timed out" in issue for issue in row["issues"])


def test_page_row_reads_schema_types_from_current_entity_shape():
    row = _page_row(
        "https://example.com/p",
        {
            "schema_check": {
                "ok": True,
                "entities": [
                    {"path": "$[0]", "types": ["WebPage"], "errors": [], "warnings": []},
                    {"path": "$[1]", "types": ["Organization"], "errors": [], "warnings": []},
                    {"path": "$[2]", "types": ["WebPage"], "errors": [], "warnings": []},
                ],
            }
        },
    )
    assert row["schema_types"] == "Organization, WebPage"


# ── input boundaries ─────────────────────────────────────────────────────────


def test_bad_input_is_rejected_before_any_network_call():
    assert audit_site("")["ok"] is False
    assert audit_site("not a URL")["ok"] is False
    assert audit_site("https://example.com/", limit="many")["ok"] is False


# ── reports ──────────────────────────────────────────────────────────────────

DOCUMENT = {
    "ok": True,
    "schema": SCHEMA,
    "url": "https://example.com/",
    "domain": "example.com",
    "generated_at": "2026-08-13T00:00:00+00:00",
    "site": {
        "tech_detect": {
            "technologies": [
                {"category": "cms", "name": "Bitrix CMS", "evidence": "found /bitrix/ in HTML"}
            ]
        },
        "domain_profile": {
            "registration": {"registrar": "RU-CENTER", "created": "2020-01-01", "age_years": 6.6}
        },
    },
    "pages": [
        {
            "url": "https://example.com/a",
            "status": 200,
            "title": "A",
            "title_length": 1,
            "description_length": 0,
            "h1": "A",
            "canonical": "https://example.com/a",
            "words": 500,
            "schema_types": "Product",
            "schema_errors": 0,
            "social_missing": 2,
            "issues": [],
        }
    ],
    "findings": [
        {"source": "render_check", "severity": "critical", "text": "empty page"},
        {"source": "cdn_check", "severity": "notice", "text": "brotli is not enabled"},
    ],
    "summary": {
        "pages_checked": 1,
        "findings_total": 2,
        "findings_by_severity": {"critical": 1, "warning": 0, "notice": 1},
        "tools_run": ["cdn_check"],
        "tools_failed": [{"tool": "log_analyze", "error": "file is unavailable"}],
        "severity_note": "severity is assigned by the aggregation rules",
    },
}


def test_every_format_produces_a_file(tmp_path):
    for fmt in FORMATS:
        result = build_report(DOCUMENT, fmt=fmt, path=str(tmp_path / f"r.{fmt}"))
        assert result["ok"] is True, f"{fmt}: {result.get('error')}"
        assert result["bytes"] > 0, f"{fmt}: empty output file"


def test_csv_writes_pages_as_a_second_file(tmp_path):
    """Findings and pages are separate entities and require separate tables."""
    target = tmp_path / "r.csv"
    build_report(DOCUMENT, fmt="csv", path=str(target))
    assert target.exists() and target.with_suffix(".pages.csv").exists()


def test_csv_is_written_with_bom_for_excel(tmp_path):
    """The BOM lets Excel detect UTF-8 correctly for multilingual report data."""
    target = tmp_path / "r.csv"
    build_report(DOCUMENT, fmt="csv", path=str(target))
    assert target.read_bytes().startswith(b"\xef\xbb\xbf")


def test_markdown_keeps_the_failed_tools_visible(tmp_path):
    target = tmp_path / "r.md"
    build_report(DOCUMENT, fmt="md", path=str(target))
    text = target.read_text(encoding="utf-8")
    assert "Unavailable checks" in text and "log_analyze" in text
    assert "Critical" in text


def test_markdown_shows_partial_sitemap_root_evidence(tmp_path):
    """One good root and one broken root (#310) must not vanish from the report the
    same way a fully clean, all-successful multi-root aggregate would."""

    def fake_sitemap_crawl(url: str | None = None, **_kw: object) -> dict:
        if url == FIRST_SITEMAP:
            return {"ok": True, "urls": [{"loc": "https://example.test/page-a"}]}
        return {"ok": True, "urls": [], "errors": [{"url": url, "error": "404"}]}

    result = _multi_root_audit(fake_sitemap_crawl)
    target = tmp_path / "r.md"
    build_report(result, fmt="md", path=str(target))
    text = target.read_text(encoding="utf-8")
    assert SECOND_SITEMAP in text


def test_excel_has_the_four_sheets(tmp_path):
    from openpyxl import load_workbook

    target = tmp_path / "r.xlsx"
    build_report(DOCUMENT, fmt="xlsx", path=str(target))
    wb = load_workbook(target)
    assert wb.sheetnames == ["Summary", "Findings", "Pages", "Technologies"]


def test_unknown_format_is_data_not_a_crash():
    result = build_report(DOCUMENT, fmt="pdf")
    assert result["ok"] is False and "pdf" in result["error"]


def test_report_can_be_built_from_a_json_path(tmp_path):
    source = tmp_path / "audit.json"
    source.write_text(json.dumps(DOCUMENT, ensure_ascii=False), encoding="utf-8")
    result = build_report(str(source), fmt="md", path=str(tmp_path / "out.md"))
    assert result["ok"] is True and result["findings"] == 2


def test_missing_audit_file_is_reported_clearly():
    result = build_report("/nope/audit.json", fmt="md")
    assert result["ok"] is False and "audit file not found" in result["error"]


def test_empty_audit_still_produces_a_report(tmp_path):
    """An audit with no findings is a valid short report, not an error."""
    result = build_report(
        {"schema": SCHEMA, "domain": "example.com", "findings": [], "pages": [], "summary": {}},
        fmt="md",
        path=str(tmp_path / "e.md"),
    )
    assert result["ok"] is True and result["findings"] == 0


def test_excel_sheets_have_no_blank_row_under_the_header(tmp_path):
    """Header styling must not materialize a blank row before the first record."""
    from openpyxl import load_workbook

    target = tmp_path / "r.xlsx"
    build_report(DOCUMENT, fmt="xlsx", path=str(target))
    wb = load_workbook(target)
    for name, expected in (
        ("Findings", len(DOCUMENT["findings"])),
        ("Pages", len(DOCUMENT["pages"])),
    ):
        ws = wb[name]
        assert ws.max_row - 1 == expected, f"{name}: unexpected rows below the header"
        first = next(ws.iter_rows(min_row=2, values_only=True))
        assert any(v is not None for v in first), f"{name}: first data row is empty"


def test_formula_leading_titles_are_neutralized_in_xlsx(tmp_path):
    """A crawled page's own title must not become a live spreadsheet formula (#153)."""
    import copy

    from openpyxl import load_workbook

    for lead in ("=", "+", "-", "@"):
        doc = copy.deepcopy(DOCUMENT)
        doc["pages"][0]["title"] = f'{lead}HYPERLINK("http://evil.example/steal","click")'
        doc["findings"][0]["text"] = f"{lead}cmd|' /C calc'!A0"
        target = tmp_path / f"r-{ord(lead)}.xlsx"
        build_report(doc, fmt="xlsx", path=str(target))
        wb = load_workbook(target)
        title_cell = wb["Pages"]["C2"]
        finding_cell = wb["Findings"]["D2"]
        assert title_cell.data_type == "s", f"lead {lead!r}: title became a live formula"
        assert finding_cell.data_type == "s", f"lead {lead!r}: finding text became a live formula"
        assert title_cell.value == "'" + doc["pages"][0]["title"]


def test_formula_leading_titles_are_neutralized_in_csv(tmp_path):
    """The CSV field must not begin with a formula-leading character either (#153)."""
    import copy
    import csv

    for lead in ("=", "+", "-", "@"):
        doc = copy.deepcopy(DOCUMENT)
        doc["pages"][0]["title"] = f'{lead}HYPERLINK("http://evil.example/steal","click")'
        target = tmp_path / f"r-{ord(lead)}.csv"
        build_report(doc, fmt="csv", path=str(target))
        with target.with_suffix(".pages.csv").open(encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh, delimiter=";"))
        title_field = rows[1][2]  # columns: url, status, title, ...
        assert not title_field.startswith(lead), f"lead {lead!r} reached the CSV cell unescaped"
        assert title_field.startswith("'")


def test_ordinary_titles_are_written_byte_for_byte_unchanged(tmp_path):
    """Titles/finding text with no formula-leading character must pass through untouched."""
    from openpyxl import load_workbook

    target = tmp_path / "r.xlsx"
    build_report(DOCUMENT, fmt="xlsx", path=str(target))
    wb = load_workbook(target)
    assert wb["Pages"]["C2"].value == DOCUMENT["pages"][0]["title"]
    assert wb["Findings"]["D2"].value == DOCUMENT["findings"][0]["text"]


def test_missing_key_security_headers_are_warnings_not_notices():
    """Missing HSTS and CSP must remain actionable warnings, not generic notices."""
    assert classify("missing strict-transport-security — enforce HTTPS") == "warning"
    assert classify("missing content-security-policy — restrict script sources") == "warning"


# ── SF Analyzer audit.json -> report contract (#220, #225) ──────────────────

SF_AUDIT = {
    "schema_version": "2.0",
    "tool": {"name": "SF Analyzer", "version": "3.0"},
    "run": {"project": "example.test", "generated_at": "2026-09-05T00:00:00Z"},
    "summary": {
        "totals": {"urls_crawled": 1, "issues_total": 1},
        "by_severity": {"critical": 1, "warning": 0, "notice": 0},
        "by_check": {"BROKEN_INTERNAL_LINK": 1},
        "health_score": 80,
    },
    "issues": [
        {
            "id": "ISSUE-000001",
            "check": "BROKEN_INTERNAL_LINK",
            "severity": "critical",
            "source": "inlinks:Client Error (4xx) Inlinks",
            "message": "Internal link points to a 4xx URL",
            "target_url": "https://example.test/dead",
            "status_code": 404,
            "occurrences_count": 2,
            "fix_hint": "Replace the shared footer link.",
            "details": {"marker": "DETAIL-MUST-SURVIVE"},
            "locations": [
                {
                    "source_url": "https://example.test/source-a",
                    "anchor": "Old page",
                    "link_position": "Footer",
                    "link_path": "/html/body/footer/a[1]",
                }
            ],
        }
    ],
    "pages": [
        {
            "url": "https://example.test/",
            "status_code": 200,
            "metrics": {
                "title": "Example",
                "title_length": 7,
                "desc_length": 159,
                "h1": ["Example"],
                "canonical": "https://example.test/",
                "word_count": 500,
            },
        }
    ],
    "groups": [],
}


def test_sf_audit_normalization_preserves_issue_evidence():
    """#220: check/status_code/occurrences_count/locations/fix_hint/details must survive.

    The SF adapter used to keep only severity/source/url/text, which discarded
    every field a BROKEN_INTERNAL_LINK finding needs for the developer handoff
    docs/scenarios/broken-pages.md promises.
    """
    from seohead.reports import _normalize_sf_audit

    finding = _normalize_sf_audit(SF_AUDIT)["findings"][0]
    assert finding["check"] == "BROKEN_INTERNAL_LINK"
    assert finding["status_code"] == 404
    assert finding["occurrences_count"] == 2
    assert finding["fix_hint"] == "Replace the shared footer link."
    assert finding["details"] == {"marker": "DETAIL-MUST-SURVIVE"}
    assert finding["locations"] == SF_AUDIT["issues"][0]["locations"]


def test_xlsx_findings_sheet_carries_located_evidence(tmp_path):
    """#220: the documented developer handoff (xlsx) must show where a broken link lives."""
    from openpyxl import load_workbook

    target = tmp_path / "sf.xlsx"
    result = build_report(SF_AUDIT, fmt="xlsx", path=str(target))
    assert result["ok"], result
    ws = load_workbook(target)["Findings"]
    headers = [cell.value for cell in ws[1]]
    row = dict(zip(headers, [cell.value for cell in ws[2]], strict=True))
    assert row["Check"] == "BROKEN_INTERNAL_LINK"
    assert row["Status"] == 404
    assert row["Occurrences"] == 2
    assert row["Fix Hint"] == "Replace the shared footer link."
    assert "https://example.test/source-a" in row["Locations"]
    assert "Old page" in row["Locations"]
    assert "Footer" in row["Locations"]
    assert "/html/body/footer/a[1]" in row["Locations"]


def test_csv_findings_carries_located_evidence(tmp_path):
    """#220: the tracker-import CSV must carry the same evidence as the xlsx handoff."""
    import csv

    target = tmp_path / "sf.csv"
    build_report(SF_AUDIT, fmt="csv", path=str(target))
    with target.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh, delimiter=";"))
    header, row = rows[0], dict(zip(rows[0], rows[1], strict=True))
    assert "Locations" in header and "Fix Hint" in header
    assert row["Check"] == "BROKEN_INTERNAL_LINK"
    assert row["Status"] == "404"
    assert row["Occurrences"] == "2"
    assert "https://example.test/source-a" in row["Locations"]


def test_xlsx_pages_sheet_has_description_length(tmp_path):
    """#225: the XLSX Pages worksheet must not drop description_length."""
    from openpyxl import load_workbook

    target = tmp_path / "sf.xlsx"
    build_report(SF_AUDIT, fmt="xlsx", path=str(target))
    ws = load_workbook(target)["Pages"]
    headers = [cell.value for cell in ws[1]]
    assert "Description Length" in headers
    row = dict(zip(headers, [cell.value for cell in ws[2]], strict=True))
    assert row["Description Length"] == 159
    assert classify("missing permissions-policy — camera access") == "notice"
